import os
import re
import openpyxl
import time
import json
from datetime import datetime
import collections
import logging
logger = logging.getLogger('utility_to_osm.ssr2.generate_webpage')

import utility_to_osm.argparse_util as argparse_util
from utility_to_osm.kommunenummer import kommunenummer, kommune_fylke
import utility_to_osm.file_util as file_util
from utility_to_osm.file_util import open_utf8
from utility_to_osm.generate_html_history_chart import render_history_chart
from utility_to_osm import osmapis

from jinja2 import Template
import humanize

from ssr2_osm import overpass_stedsnr_in_kommunenr
from osmapis_stedsnr import OSMstedsnr

FAST_MODE = False                # For development, skip some slow parts, fixme: command-line argument

link_template = u'<a href="{href}" title="{title}">{text}</a>'
progress_template = '''
<meter title="{per:.2f}% = {value}/{max} nodes with ssr:stedsnr in OSM, last checked {overpass_checked}" 
style="width:100%" value="{value}" min="{min}" max="{max}" optimum="{max}">{per} %</meter>
'''.strip()
progress_template_circle = '''
<span title="{per:.2f}% = {value}/{max} nodes with ssr:stedsnr in OSM, last checked {overpass_checked}" 
class="dot" style="background-color: {color};"></span>
'''.strip()

footer = """
<!-- FOOTER  -->
<div id="footer_wrap" class="outer">
  <footer class="inner">
    <p class="copyright">
      This page is maintained by <a href="https://github.com/obtitus">obtitus</a>
    </p>
    <p class="copyright">
      OSM data &copy;<a href="http://openstreetmap.org">OpenStreetMap</a> contributors,
      all osm data licensed under
      <a href=https://opendatacommons.org/licenses/odbl/>ODbL</a>
    </p>
    <p class="copyright">
      All place data extracted from SSR &copy;<a href="https://kartverket.no">kartverket.no</a> under 
      <a href="https://creativecommons.org/licenses/by/3.0/no/">CC BY 3.0</a>
    </p>
</footer>
</div>
"""

header = """
<!-- HEADER -->
<div id="header_wrap" class="outer">
  <header class="inner">
    <a id="forkme_banner" href="https://github.com/obtitus/ssr2_to_osm">View on GitHub</a>
    <h1 id="project_title">SSR2 import to OpenStreetMap.org</h1>
    <h2 id="project_tagline">Data files for importing Norwegian placenames from Kartverket, ssr2, into OpenStreetMap</h2>
  </header>
</div>
"""

info = """
The table below contains <a href=https://josm.openstreetmap.de>JOSM</a> files for import to OSM, each municipality (kommune)
has to following files:
<ol>
<li> A "Dataset for import" file containing all of the data, including some additional
raw SSR tags for filtering and debugging that should not be uploaded to OSM. (Do not upload ssr:type, ssr:gruppe, ssr:hovedgruppe nor ssr:date)</li>
<li> A "Excerpts for import" containing subsets of the data, ready for import.</li>
<li> A "Excluded from import" column with any data that is either missing a 
<a href=https://wiki.openstreetmap.org/wiki/Key:name>name=*</a>
tag (but typically contains either 
<a href=https://wiki.openstreetmap.org/wiki/Key:old_name>old_name=*</a> or
<a href=https://wiki.openstreetmap.org/wiki/Key:loc_name>loc_name=*</a>).
The column also contain data that lacks a translation from SSR to OSM tags (see <a href=https://drive.google.com/open?id=1krf8NESSyyObpcV8TPUHInUCYiepZ6-m>tagging table</a>), typically excluded since these
are coved by separate imports. 
Data from this column should in general not be imported.</li>
</ol>
See the <a href=http://wiki.openstreetmap.org/wiki/No:Import_av_stedsnavn_fra_SSR2>import wiki</a> for further details.
"""

# <li> Raw data from Kartverket (SSR) in the rightmost column. </li>


# 
# Statistics gathering
# 
global ssr_type_tags
global ssr_type_freq
ssr_type_tags = dict() # Key is the 'ssr:type' while the value is a defaultdict counter
ssr_type_freq = collections.defaultdict(int) # Key is the 'ssr:type', value is the number of occurrences in ssr


def progress_to_color(per):
    """My own wierd colormap going to the 'Greenery' color #92B558 as the maximum
    """
    assert 0 <= per <= 100, 'expected percentage between 0 and 100, got %s' % (per)

    red_max = 0x92
    red = min(red_max, 10*int(red_max*per/100.))
    green_max = 0xb5
    green = min(green_max, int(2+green_max*per/100.))
    blue_max = 0x58
    blue = min(blue_max, 10*int(blue_max*per/100.))
    return '#%02x%02x%02x' % (red, green, blue)

# Testing colormap:
# for per in range(0, 100):
#     color = progress_to_color(per)
#     info += progress_template_circle.format(value=per, max=100, overpass_checked='', color=color)

def create_text(filename, f, overpass=None, stedsnr_duplicates=None):
    overpass_osm_mapping = dict()
    if overpass is None:
        overpass = []
    
    if stedsnr_duplicates is None:
        stedsnr_duplicates = set()
    
    N_nodes = -1
    
    if f.endswith('.osm'):
        if re.match('\d\d\d\d-', f):
            f = f[5:]
            f = f.replace('-', ' ')
            f = f.replace('tagged', '')
            f = f.replace('offentligAdministrasjon', 'offentlig')
            f = f.replace(' .osm', '.osm')
            f = f.strip()
            if f == '.osm':
                f = 'all.osm'

        if 'clean' in filename:
            f = 'clean-' + f

        osm = list()
        if not(FAST_MODE):
            content = file_util.read_file(filename)
            osm = OSMstedsnr.from_xml(content)

            # statistics
            for ssr in osm:
                ssr_type = ssr.tags.get('ssr:type')
                if ssr_type is not None:
                    ssr_type_freq[ssr_type] += 1
            
            # Add duplicates, if any:
            for item in osm.stedsnr_duplicates:
                #if isinstance(item, osmapis.Node)
                if item.xml_tag == 'node':
                    stedsnr_duplicates.add(item)
            
            # Count objects in overpass that is also in osm file
            if overpass_osm_mapping is not None:
                for item in overpass:
                    if 'ssr:stedsnr' in item.tags:
                        stedsnr = item.tags['ssr:stedsnr']
                        if stedsnr in osm.stedsnr:
                            #overpass_osm_mapping.add(item.tags['ssr:stedsnr'])
                            overpass_osm_mapping[item.tags['ssr:stedsnr']] = (osm.stedsnr[stedsnr][0], item)
                        #N_overpass += 1
                    #else:
                    # fixme: Warning only when reading the 'full' file
                    #    logger.warning('Found ssr:stedsnr = %s in OSM but not in kommune')
        
        N_nodes = len(osm)
        N_nodes_str = '%s node' % N_nodes
        if N_nodes > 1:
            N_nodes_str += 's'

        # N_bytes = os.path.getsize(filename)
        # N_bytes_str = humanize.naturalsize(N_bytes, format='%d')

        f = '%s (%s)' % (f, N_nodes_str)
    else:
        pass

    N_overpass = -1
    if overpass_osm_mapping is not None:
        N_overpass = len(overpass_osm_mapping)


    # Gather statistics from overpass_osm_mapping
    global ssr_type_tags # 
    #print(overpass_osm_mapping)
    for stedsnr, (ssr, overpass) in overpass_osm_mapping.items():
        #print('%s:\n%s\n%s' % (stedsnr, ssr.tags, overpass.tags))
        if 'ssr:type' not in ssr.tags:
            continue

        ssr_type = ssr.tags['ssr:type']
        if ssr_type not in ssr_type_tags:
            ssr_type_tags[ssr_type] = collections.defaultdict(int)

        for key, value in overpass.tags.items():
            if key not in ('wikidata', 'population', 'ele', 'source', 'fixme', 'operator', 'website', 'ref', 'maxspeed', 'phone') \
               and not(key.endswith(('name', 'date'))) and not(key.startswith(('name:', 'ssr:', 'seamark:light', 'wikipedia', 'addr'))):
                ssr_type_tags[ssr_type]['%s = %s' % (key, value)] += 1

        #print(ssr_type_tags)
        
    return f, N_nodes, N_overpass

def write_template(template_input, template_output, **template_kwargs):
    with open_utf8(template_input) as f:
        template = Template(f.read())

    template_kwargs['header'] = template_kwargs.pop('header', header)
    template_kwargs['footer'] = template_kwargs.pop('footer', footer)
    template_kwargs['chart'] = template_kwargs.pop('chart', footer)
    page = template.render(**template_kwargs)

    with open_utf8(template_output, 'w') as output:
        output.write(page)

def create_row(kommune_nr, folder, cache_dir, stedsnr_duplicates,
               kommune_nr2name, kommuneNr_2_fylke,
               fylker,
               N_overpass_total, N_ssr_total):
    kommune_name = None
    kommune_nr_int = None
    fylke_nr = 0
    try:
        kommune_nr_int = int(kommune_nr)                
        kommune_name = kommune_nr2name[kommune_nr_int]
        if not(kommune_name.startswith(u'Longyearbyen')):
            kommune_name += ' kommune'

    except KeyError as e:
        logger.warning('Could not translate kommune_nr = %s to a name. Skipping', kommune_nr)
        #kommune_name = 'ukjent'
        kommune_nr_int = None
    except ValueError as e:
        if kommune_nr == 'ZZ':
            kommune_name = 'Outside mainland'
        else:
            raise ValueError(e)

    try:
        fylke_name, fylke_nr = kommuneNr_2_fylke[int(kommune_nr)]
        t = (fylke_name, fylke_nr)
        if t not in fylker:
            fylker.append(t)

        if not(fylke_name.startswith('Svalbard')):
            fylke_name += ' fylke'

    except KeyError as e:
        logger.warning('Could not translate kommune_nr = %s to a fylke-name. Skipping', kommune_nr)
        #fylke_name = 'ukjent'
        #return None, None, None
        fylke_nr = None
    except ValueError:
        if kommune_nr == 'ZZ':
            fylke_name = ''
        else:
            raise ValueError(e)

    overpass = None
    N_overpass = -1
    overpass_checked_date = ''
    if not(FAST_MODE) and kommune_nr_int is not None:
        cache_filename = os.path.join(folder, 'overpass', '%s-overpass-stedsnr.osm' % kommune_nr)
        file_util.create_dirname(cache_filename)
        #cache_filename = os.path.join(folder, '%s-osmStedsnr.osm' % kommune_nr)
        try:
            overpass = overpass_stedsnr_in_kommunenr(kommune_nr_int,
                                                     cache_filename=cache_filename,
                                                     cache_dir=cache_dir)

            timestamp = os.path.getmtime(cache_filename)
            overpass_checked_datetime = datetime.fromtimestamp(timestamp)
            overpass_checked_date = overpass_checked_datetime.strftime('%Y-%m-%d %H:%M')

        except KeyError as e:
            logger.error('Could not translate kommune_nr = %s to a osm-relation-id. Skipping overpass call: %s' % (kommune_nr_int, e))
            overpass_checked_date = 'Could not translate kommune_nr = %s to a osm-relation-id, overpass not called' % kommune_nr_int

        N_overpass = len(overpass)

    row = list()
    dataset_for_import = [] # expecting a single entry here
    excluded_from_import = []
    excerpts_for_import = []
    raw_data = []
    log = []
    if kommune_nr_int is not None:
        for root, dirs, files in os.walk(folder):
            for f in files:
                f = f.decode('utf8')
                if f.endswith(('.osm', '.xml', '.log', '.gml')):
                    filename = os.path.join(root, f)
                    text, N_ssr, N_overpass = create_text(filename, f, overpass=overpass,
                                                          stedsnr_duplicates=stedsnr_duplicates)

                    if N_ssr != 0:
                        per = (100.*N_overpass)/N_ssr

                    href = filename.replace('html/', '')
                    url = link_template.format(href=href,
                                               title=filename,
                                               text=text)
                    if root.endswith('clean'):
                        progress_str = ''
                        if N_ssr != -1 and N_overpass != -1:
                            color = progress_to_color(per)
                            progress = progress_template_circle.format(value=N_overpass, max=N_ssr, per=per,
                                                                       overpass_checked=overpass_checked_date,
                                                                       color=color)
                            progress_str = '&thinsp;%s' % progress

                        excerpts_for_import.append(url + progress_str)
                    elif f.endswith('%s.osm' % kommune_nr):
                        dataset_for_import.append(url)
                        if N_ssr != -1 and N_overpass != -1:
                            progress = progress_template.format(value=N_overpass, min=0, max=N_ssr, per=per,
                                                                overpass_checked=overpass_checked_date)
                            dataset_for_import.append(progress)
                            N_ssr_total += N_ssr
                            N_overpass_total += N_overpass
                    elif 'NoName' in f or 'NoTags' in f:
                        progress_str = ''
                        if N_ssr != -1 and N_overpass != -1:
                            color = progress_to_color(per)
                            progress = progress_template_circle.format(value=N_overpass, max=N_ssr, per=per,
                                                                       overpass_checked=overpass_checked_date,
                                                                       color=color)
                            progress_str = '&thinsp;%s' % progress

                        excluded_from_import.append(url + progress_str)
                    elif f.endswith(('.xml', '.gml')):
                        raw_data.append(url)
                    elif f.endswith('.log'):
                        log.append(url)
                    else:
                        pass # ignore

        row.append(fylke_nr)
        row.append("%s" % fylke_name)

        kommune_cell = list(log)
        kommune_cell.insert(0, "%s %s" % (kommune_nr, kommune_name))
        row.append(kommune_cell)            
        row.append(dataset_for_import)
        row.append(excerpts_for_import)            
        row.append(excluded_from_import)
        #row.append(raw_data) # skip raw data to reduce storage requirements
    
    return row, N_overpass_total, N_ssr_total

        
def create_main_table(data_dir='output', cache_dir='data'):
    N_overpass_total = 0
    N_ssr_total = 0
    table = list()
    stedsnr_duplicates = set()
    last_update = ''
    kommune_nr2name, kommune_name2nr = kommunenummer(cache_dir=cache_dir)
    kommuneNr_2_fylke = kommune_fylke(cache_dir=cache_dir)
    fylker = list()
    
    for kommune_nr in os.listdir(data_dir):
        # if kommune_nr == '0213':
        #     break
        
        kommune_nr_int = None
        folder = os.path.join(data_dir, kommune_nr)
        if os.path.isdir(folder):
            # mod_time = file_util.folder_modification(folder)
            # if mod_time != -1:
            #     ??time.ctime(mod_time)
                
            #print time.ctime(mod_time)
            # fixme: class instead of this mess
            row, N_overpass_total, N_ssr_total = create_row(kommune_nr, folder, cache_dir=cache_dir,
                                                            kommune_nr2name=kommune_nr2name,
                                                            kommuneNr_2_fylke=kommuneNr_2_fylke,
                                                            stedsnr_duplicates=stedsnr_duplicates,
                                                            N_ssr_total=N_ssr_total, fylker=fylker,
                                                            N_overpass_total=N_overpass_total)
            
            if row is None:
                continue
            
            table.append(row)

            # Fixme: use max of mod_time? Is missleading as it does not imply a fetch from ssr.
            last_update_stamp = os.path.getmtime(folder)
            last_update_datetime = datetime.fromtimestamp(last_update_stamp)
            last_update = last_update_datetime.strftime('%Y-%m-%d') # Note: date is now set by whatever row is 'last'

            # if len(table) > 10:
            #     break
            # if len(stedsnr_duplicates) > 2:
            #     break

    return table, fylker, last_update, stedsnr_duplicates, N_overpass_total, N_ssr_total

def main(data_dir='ssr2_to_osm_data/data/', root_output='ssr2_to_osm_data', template='ssr2_to_osm_data/templates/index_template.html'):
    output_filename = os.path.join(root_output, 'index.html')
    output_filename_hist = os.path.join(root_output, 'history.csv')
    output_filename_stat_json = os.path.join(root_output, 'stat.json')
    output_filename_stat_excel = os.path.join(root_output, 'stat.xlsx')
    
    table, fylker, last_update, stedsnr_duplicates, N_overpass_total, N_ssr_total = create_main_table(data_dir, cache_dir=data_dir)
    errors = ''
    if len(stedsnr_duplicates) != 0:
        xml = osmapis.OSM()
        for item in stedsnr_duplicates:
            xml.add(item)
        filename_dup = 'duplicates.xml'
        xml.save(os.path.join(root_output, filename_dup))

        link = link_template.format(href=filename_dup,
                                    title=filename_dup,
                                    text=filename_dup)
        s = ''
        if len(stedsnr_duplicates) != 1:
            s = 's'
        errors = '''<div class="error">Duplicate element%s found, %s node%s with duplicated 
        ssr:nsrid found in osm, please see %s</div>''' % (s, len(stedsnr_duplicates), s, link)


    # dump progress to csv
    today = datetime.utcnow()
    td = (today - datetime(1970, 1, 1))
    td_s = td.total_seconds()
    with open(output_filename_hist, 'a') as f:
        f.write('%s,%s,%s\n' % (td_s, N_ssr_total, N_overpass_total))

    # read progress from csv
    chart = render_history_chart(root_output,
                                 header=['Number of noder/ways with ssr:stedsnr in OSM', ''],
                                 title='Total number of nodes',)

    # Fill out the template
    write_template(template, output_filename, table=table, info=info,
                   fylker=fylker, errors=errors,
                   last_update=last_update,
                   chart=chart)

    # Statistics
    for key in ssr_type_tags.keys():
        # max count
        max_count = 0
        for tag, count in ssr_type_tags[key].items():
            max_count = max(max_count, count)
        # remove tags with only 1 value and remove those with significant less than max count
        for tag, count in ssr_type_tags[key].items():
            if count == 1 or count < max_count//25:
                del ssr_type_tags[key][tag]

    with open(output_filename_stat_json, 'w') as f:
        json.dump(ssr_type_tags, f)

    from openpyxl.styles import Alignment
    wb = openpyxl.load_workbook(os.path.join('data', 'Tagging tabell SSR2.xlsx'))
    ws = wb['Taggetabell'] #wb.create_sheet()
    ws.page_setup.fitToWidth = 1
    header = next(ws.rows)
    ix_navnetype = 0
    ix_ssr2_count = 0
    ix_tags = 0
    for h in header:
        if h.value == 'SSR2 navnetype':
            ix_navnetype = h.col_idx
        elif h.value == 'SSR2 antall':
            ix_ssr2_count = h.col_idx
        elif h.value == 'Tags brukt i OSM, antall':
            ix_tags = h.col_idx

    #ws.append(['SSR2 navnetype', 'SSR2 freq', 'Tag usage in OSM'])
    for key in ssr_type_tags.keys():
        tag_cell = []
        for tag, count in ssr_type_tags[key].items():
            tag_cell.append('"%s", %s' % (tag, count))
        tag_cell_str = '\n'.join(tag_cell)

        freq_in_ssr = ssr_type_freq.get(key, 0)

        # Data ready, lets find correct row to insert into
        for row_ix in range(ws.max_row):
            if ws.cell(column=ix_navnetype, row=row_ix+1).value == key:
                break
        else:
            raise ValueError('Could not find "%s" in column %s' % (key, ix_navnetype))

        #ws.append([key, freq_in_ssr, cell])
        ws.cell(column=ix_ssr2_count, row=row_ix+1).value = freq_in_ssr
        
        cell = ws.cell(column=ix_tags, row=row_ix+1)
        cell.value = tag_cell_str
        cell.alignment = Alignment(wrapText=True)
        

    wb.save(output_filename_stat_excel)
    
if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='')
    argparse_util.add_verbosity(parser, default=logging.INFO)
    
    args = parser.parse_args()
    logging.basicConfig(level=args.loglevel)

    main()
